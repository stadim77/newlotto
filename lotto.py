# import pandas as pd
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ---import turn ---#

filename = ['lotto_2018.xlsx', 'lotto_2019.xlsx', 'lotto_2020.xlsx', 'lotto_2021.xlsx']

try:
    print('Enter an integer from 0 to 3 :')
    ans = int(input())
except ValueError:
    print('Not a valid value !!!')
wb = load_workbook(filename[ans])
### process to remove None rows from data ###
data = pd.read_excel(filename[ans])
df1 = data.dropna()
# ---print(df1.shape) --> to find elements dimensions
print(wb.sheetnames)

### ENTER THE LOTTO NUMBERS ###
nums = []
i = 1
while i < 7:
    print("Enter the number that you want to search.")
    try:
        ans = int(input())
        if not 1 <= ans <= 49:
            raise ValueError()
    except ValueError:
        print('Please enter a valid number!!!')
    else:
        nums.append(ans)
        i += 1

print(f"You entered the numbers : {nums} ")

sheet = wb.active
# ---print(sheet.max_row) # prints the rows of sheet (draws of lotto)
# print(sheet.title)
s = list(sheet.rows)


# ---print(len(s))
# ---print(list(sheet.rows)[2])

def print_values(rowval, ran):
    ''' get element values from specific rows and columns'''
    lista = []
    for item in range(1, ran):
        lista.append(sheet.cell(rowval, item).value)
        # print(sheet.cell(rowval,item).value)
    return print(lista[2:])  # prints num elements from list


def get_values(rowval, ran=10):
    ''' get element values from specific rows and columns (until 10, to get the numbers)'''
    lista = []
    for item in range(1, ran):
        lista.append(sheet.cell(rowval, item).value)
        # print(sheet.cell(rowval,item).value)
    return lista


### TESTING ###
###for i in range(5,8):
###    print_values(i,10) # prints lotto numbers of each row

# --- a = get_values(5,10) # FIRST ARGUMENT MUST BE > 5 BECAUSE OF THE TITLES 1-4 ROWS
b = get_values(7, 12)


# --- print(a)
# --- print(b)
# --- print(b[2:])
# print(type(b[3]))

def luck():
    ''' returns the victories'''
    for i in range(5, df1.shape[0]):
        counts = 0
        if get_values(i, 12)[-1] == '1' or get_values(i, 12)[-1] == '2':
            print(f'Perfect match : {get_values(i)[:2]}')
            # print(f'Numbers : {get_values(i)[2:]}')
            counts += 1
        else:
            pass


print(luck())


def cons_nums():
    '''returns the 6 numbers from a lotto year and the bonus 7th number'''
    numbers = []
    bonus = []
    for i in range(5, df1.shape[0]):  # df1.shape --> get the max valid rows
        numbers.append(get_values(i)[2:-1])
        bonus.append(get_values(i)[-1])
    # [elem for elem in numbers if elem is not None]
    return numbers


print('#####')
# ---print(cons_nums())

### CONVERT TO NUMPY ARRAY OF INTEGERS ###
np_array = np.array(cons_nums())
print(np_array.shape)
# ---np_array = np_array[np_array != None]
draws = np_array.astype(np.int64)
# ---print(draws) # the rows and columns of array
df = pd.DataFrame(draws, columns=['1st', '2nd', '3rd', '4th', '5th', '6th'])
# ---print(df)
### GET THE TOP 3 VALUES FOR EACH POSITION ###
print(df['1st'].value_counts()[:3])
print(df['2nd'].value_counts()[:3])
print(df['3rd'].value_counts()[:3])
print(df['4th'].value_counts()[:3])
print(df['5th'].value_counts()[:3])
print(df['6th'].value_counts()[:3])


### GET THE SUM OF EACH ROW ###
###---df['sum'] = df.sum(axis=1)
###---df['mean'] = df.mean(axis=1)
# filtering rows by 'sum' conditions
###---print(df[df['sum'] > 190])
###---print(df[df['sum'] < 95])
# ---print(df[(df['sum'] > 100) & (df['sum'] < 150)])
# ---print(df[(df['sum'] > 100) & (df['mean'] < 32)])

### FIND OCCURENCES IN PREVIOUS DRAWS ###
def Diff(li1, li2):
    '''returns the difference between two lists'''
    return list(set(li1) - set(li2)) + list(set(li2) - set(li1))


def cuts(dataframe, mun):
    ''' returns the number of exact matches'''
    ans = 0  # complete match
    fives = 0  # 5 numbers that match
    fours = 0  # 4 numbers that match
    threes = 0  # 3 numbers that match
    twos = 0  # 2 numbers that match
    for i in range(len(dataframe)):
        a = list(map(int, list(dataframe.iloc[i, 2:7])))
        if len(Diff(a, mun)) == 0:
            ans += 1
        elif len(Diff(a, mun)) == 2:
            fives += 1
        elif len(Diff(a, mun)) == 4:
            fours += 1
        elif len(Diff(a, mun)) == 6:
            threes += 1
        elif len(Diff(a, mun)) == 8:
            twos += 1
        else:
            pass
    # print(ans,fours,threes,twos)
    print(
        f'Complete matches : {ans},Five number matches: {fives},  Four number matches : {fours}, Three number matches : {threes} and two number matches : {twos}')
    return (ans, fives, fours, threes, twos)


cuts(df, nums)

### GET THE SUM OF EACH ROW ###
df['sum'] = df.sum(axis=1)
###---df['mean'] = df.mean(axis=1)
# filtering rows by 'sum' conditions
print(df[df['sum'] > 190])


###---print(df[df['sum'] < 95])

### create an object of LOTTO results
###l = turn.Turn(a)

###l.__repr__()

###l1 = turn.Turn(b)

###l1.__repr__()

###l2 = turn.Turn(get_values(8))

###l2.__repr__()0

# objects = list()
# for i in range(5,sheet.max_row):
#     if sheet.cell(row = i, column = 1).value is None:
#         # checks if value of cell[0] is None, and avoids it
#         pass
#     else:
#         i = turn.Turn(get_values(i))
#         i.__repr__()
#         objects.append(i.data[0])

# ---print(objects[:]) # prints the list of numbers of events
# ---print(len(objects))

# for item in objects:
#     i = 0
#     objects[int(item)] = turn.Turn(get_values(i+5))
#     print(f"{objects[int(item)]} : has the numbers {objects[int(item)].data[2]}, {objects[int(item)].data[3]}")
#     i += 1
###print(l1.__nums__[:2])

###print(l.__nums__[:2])

# ---print(df[df['1st'] == 10].index.values)
# df_mask = df['1st']==12
# print(df.iloc[df_mask])

### CREATE A FUNCTION TO GET ROW INDEXES FOR ROWS THAT CONTAIN A NUMBER ###
def icols(frame, num):
    '''returns the indexes of occurence of selected number'''
    res = []
    for column in frame:
        # res.append(frame[frame[column] == num].index.values)
        res.extend(frame[frame[column] == num].index.values)
        # print(frame[frame[column] == num].index.values)
    return res


print(icols(df, nums[0]))
print(icols(df, nums[1]))
print(icols(df, nums[2]))
print(icols(df, nums[3]))
print(icols(df, nums[4]))
print(icols(df, nums[5]))

### the occurences of num[0] ###
ad_1 = icols(df, nums[0])
print(f'the occurences of {nums[0]} is {sorted(ad_1)}')
ad_2 = icols(df, nums[1])
print(f'the occurences of {nums[1]} is {sorted(ad_2)}')
ad_3 = icols(df, nums[2])
print(f'the occurences of {nums[2]} is {sorted(ad_3)}')
ad_4 = icols(df, nums[3])
print(f'the occurences of {nums[3]} is {sorted(ad_4)}')
ad_5 = icols(df, nums[4])
print(f'the occurences of {nums[4]} is {sorted(ad_5)}')
ad_6 = icols(df, nums[5])
print(f'the occurences of {nums[5]} is {sorted(ad_6)}')

# ---print(df[df['1st'] == 10])

### RETURN THE 4s,3s,2s SUCCESSES ###
print('4 matches')
print(df.loc[df.isin(nums).sum(axis=1) == (len(nums) - 2), :])
print('3 matches')
print(df.loc[df.isin(nums).sum(axis=1) == (len(nums) - 3), :])
print('2 matches')
print(df.loc[df.isin(nums).sum(axis=1) == (len(nums) - 4), :])

# m = [nums.issubset(i) for i in df.values.tolist()]
# print(df[m])

# turn a dataframe's column to list
print(df['1st'].tolist())

t1 = df['1st'].tolist()

print(nums)
if nums[0] in t1:
    print(t1.index(nums[0]))
else:
    pass

if nums[1] in t1:
    print(t1.index(nums[1])) #TODO problem when number is not in list
else:
    pass

# find all the occurences in list
te = []
[te.append(i) for i, e in enumerate(t1) if e == nums[1]]
print(te)

### return the occurences of number in each column/position
def loc(datafr,nm):
    '''return the occurences of number in each column/position of LOTTO dataframe'''
    tel = []
    tf1 = df['1st'].tolist()
    tf2 = df['2nd'].tolist()
    tf3 = df['3rd'].tolist()
    tf4 = df['4th'].tolist()
    tf5 = df['5th'].tolist()
    tf6 = df['6th'].tolist()
    [tel.append(i) for i, e in enumerate(tf1) if e == nm]
    [tel.append(i) for i, e in enumerate(tf2) if e == nm]
    [tel.append(i) for i, e in enumerate(tf3) if e == nm]
    [tel.append(i) for i, e in enumerate(tf4) if e == nm]
    [tel.append(i) for i, e in enumerate(tf5) if e == nm]
    [tel.append(i) for i, e in enumerate(tf6) if e == nm]
    return sorted(tel)

print(f'the {nums[0]} occured in {loc(df,nums[0])} draws.')
print(f'the {nums[1]} occured in {loc(df,nums[1])} draws.')
print(f'the {nums[2]} occured in {loc(df,nums[2])} draws.')
print(f'the {nums[3]} occured in {loc(df,nums[3])} draws.')
print(f'the {nums[4]} occured in {loc(df,nums[4])} draws.')
print(f'the {nums[5]} occured in {loc(df,nums[5])} draws.')

### return the common occurences of two numbers

def common(nm1,nm2):
    '''find the draws where the 2 numbers occured'''
    ansc = []
    for i in loc(df,nm1):
        if i in loc(df,nm2):
            ansc.append(i)
        else:
            pass
    return ansc

print(f'{nums[0]} and {nums[1]} occured in {common(nums[0],nums[1])} draw.')
print(f'{nums[0]} and {nums[2]} occured in {common(nums[0],nums[2])} draw.')
print(f'{nums[0]} and {nums[3]} occured in {common(nums[0],nums[3])} draw.')
print(f'{nums[0]} and {nums[4]} occured in {common(nums[0],nums[4])} draw.')
print(f'{nums[0]} and {nums[5]} occured in {common(nums[0],nums[5])} draw.')

print('---------------------')
print(f'{nums[1]} and {nums[2]} occured in {common(nums[1],nums[2])} draw.')
print(f'{nums[1]} and {nums[3]} occured in {common(nums[1],nums[3])} draw.')
print(f'{nums[1]} and {nums[4]} occured in {common(nums[1],nums[4])} draw.')
print(f'{nums[1]} and {nums[5]} occured in {common(nums[1],nums[5])} draw.')

print('---------------------')
print(f'{nums[2]} and {nums[3]} occured in {common(nums[2],nums[3])} draw.')
print(f'{nums[2]} and {nums[4]} occured in {common(nums[2],nums[4])} draw.')
print(f'{nums[2]} and {nums[5]} occured in {common(nums[2],nums[5])} draw.')
print('----------------------')
print(f'{nums[3]} and {nums[4]} occured in {common(nums[3],nums[4])} draw.')
print(f'{nums[3]} and {nums[5]} occured in {common(nums[3],nums[5])} draw.')
print('----------------------')
print(f'{nums[4]} and {nums[5]} occured in {common(nums[4],nums[5])} draw.')



